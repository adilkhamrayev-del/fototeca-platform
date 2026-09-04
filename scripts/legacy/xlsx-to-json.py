#!/usr/bin/env python3
"""Converts a "Журнал заказов" export from the old XAF system
(zakaz.fototeca.kz -> Журнал заказов -> Экспорт -> Excel файл (.XLSX)) into a
flat JSON array that scripts/migrate-legacy-orders.ts can read.

Kept as a separate step (rather than parsing .xlsx directly in Node) because
the platform intentionally has no xlsx-parsing dependency, and Python +
openpyxl is a much simpler tool for a one-off/occasional export than adding
an npm package just for this.

Usage:
    python3 scripts/legacy/xlsx-to-json.py path/to/Заказы.xlsx path/to/output.json

Expected columns (as exported by the grid, in this order — the script reads
by header name, not position, so a reordered export still works):
    № заказа, Дата, Подразделение, Контрагент, Номер ячейки, Сумма заказа,
    К оплате, Сумма оплаты, Электронная почта, Исполнители, Срок сдачи,
    Данные о доставке
"""

import json
import re
import sys
from datetime import datetime

import openpyxl

# Kazakhstan mobile numbers written every possible way in the "Контрагент"
# free-text field: "+7 701 131 2265", "8 701 996 09 11", "87778475006", etc.
PHONE_RE = re.compile(r"(?:\+?7|8)[\s\-]?\(?7\d{2}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}")


def normalize_phone(raw: str) -> str | None:
    match = PHONE_RE.search(raw)
    if not match:
        return None
    digits = re.sub(r"\D", "", match.group())
    if digits.startswith("8") and len(digits) == 11:
        digits = "7" + digits[1:]
    if len(digits) != 11 or not digits.startswith("7"):
        return None
    return digits


def split_contractor(raw: str) -> tuple[str, str | None]:
    raw = (raw or "").strip()
    phone = normalize_phone(raw)
    # Strip the matched phone (and common city-prefix noise) out of the name
    # so `contractor_name` stays reasonably clean; keep the raw string as a
    # fallback if this leaves nothing useful.
    name = raw
    if phone:
        name = PHONE_RE.sub("", raw).strip(" -,")
    name = re.sub(r"\s{2,}", " ", name).strip()
    return name or raw, phone


def iso(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    src, dst = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else "" for h in next(rows)]
    col = {name: i for i, name in enumerate(header)}

    required = ["№ заказа", "Дата", "Контрагент", "Сумма заказа"]
    missing = [c for c in required if c not in col]
    if missing:
        print(f"Export is missing expected columns: {missing}. Found: {header}")
        sys.exit(1)

    def get(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in rows:
        legacy_number = get(row, "№ заказа")
        if not legacy_number:
            continue
        contractor_raw = str(get(row, "Контрагент") or "").strip()
        contractor_name, phone = split_contractor(contractor_raw)

        out.append(
            {
                "legacyNumber": str(legacy_number).strip(),
                "legacyDate": iso(get(row, "Дата")),
                "department": (get(row, "Подразделение") or None),
                "contractorRaw": contractor_raw or None,
                "contractorName": contractor_name or None,
                "phone": phone,
                "cellNumber": (str(get(row, "Номер ячейки")).strip() if get(row, "Номер ячейки") else None),
                "orderAmount": get(row, "Сумма заказа"),
                "amountToPay": get(row, "К оплате"),
                "amountPaid": get(row, "Сумма оплаты"),
                "email": (str(get(row, "Электронная почта")).strip() if get(row, "Электронная почта") else None),
                "executor": (get(row, "Исполнители") or None),
                "dueDate": iso(get(row, "Срок сдачи")),
                "deliveryInfo": (get(row, "Данные о доставке") or None),
            }
        )

    with open(dst, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=None)

    with_phone = sum(1 for r in out if r["phone"])
    with_email = sum(1 for r in out if r["email"])
    print(f"Wrote {len(out)} rows to {dst}")
    print(f"  {with_phone} rows ({100 * with_phone // max(len(out), 1)}%) have a parseable phone number")
    print(f"  {with_email} rows ({100 * with_email // max(len(out), 1)}%) have an email")


if __name__ == "__main__":
    main()
